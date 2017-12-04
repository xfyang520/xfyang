#_*_ coding:utf-8 _*_
__author__ = 'xfyang'

import openpyxl
import datetime

if __name__ == '__main__':
    # excel文件全路径
    xlPath = "E:\\xfyang\\appiumTest\\testExcel.xlsx"

    # 生成新建文件的文件名
    now = datetime.datetime.now()
    nowTime = str(now.year) + str(now.month) + str(now.day) + str(now.hour) + str(now.minute) + str(
        now.second) + "_" + str(now.microsecond)
    print(nowTime)

    # 用于读取excel
    pyxlBook = openpyxl.load_workbook(xlPath)

    print(pyxlBook.sheetnames)
    pySheet = pyxlBook.get_sheet_by_name("Sheet1")

    for i in range(1,pySheet.max_row+1):
        print(i)
        for j in pySheet[i]:
            print(j.value)

    #pySheet["A15"].value ="TestExcel014"

    pyxlBook.save("E:\\xfyang\\appiumTest\\" + nowTime + ".xlsx")  # 另存为一个新文件
    #pyxlBook.save(xlPath)   # 保存当前文件
    print("保存成功")
    pyxlBook.close()